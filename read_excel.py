from openpyxl import load_workbook

def get_env(location):
    result = {}
    wb = load_workbook(location)
    content = wb.get_sheet_by_name("ENV")
    row_num = content.max_row
    column_num = content.max_column
    for r in range(2,row_num+1):
        table = {}
        for c in range(2,column_num+1):
            if c == 2:
                table['netdata.%s.user' % content.cell(row=1, column=c)._value] = "last(value)"
            else:
                if content.cell(row=r, column=c)._value != 0:
                    table['netdata.%s.used' % content.cell(row=1, column=c)._value] = 'last(value)*100/(%s*1024)' % content.cell(row=r, column=c)._value
        result[content.cell(row=r,column=1)._value] = table
    print(result)
    return result

def get_threshold(location):
    thresholds = {}
    wb = load_workbook(location)
    content = wb.get_sheet_by_name("Threshold")
    column_num = content.max_column
    for c in range(2,column_num+1):
        td = []
        td.append(content.cell(row=2, column=c)._value)
        td.append(content.cell(row=3, column=c)._value)
        if c == 2:
            thresholds['netdata.%s.user' % content.cell(row=1,column=c)._value] = td
        thresholds['netdata.%s.used' % content.cell(row=1, column=c)._value] = td
    print thresholds
    return thresholds

def get_host(location):
    host_list = []
    wb = load_workbook(location)
    content = wb.get_sheet_by_name("ENV")
    row_num = content.max_row
    for r in range(2,row_num+1):
        host_list.append(content.cell(row=r,column=1)._value)
    print host_list
    return host_list

def get_measurement(location):
    measurement_list = {}
    wb = load_workbook(location)
    content = wb.get_sheet_by_name("ENV")
    column_num = content.max_column
    row_num = content.max_row
    for r in range(2,row_num+1):
        meas = []
        for c in range(2,column_num+1):
            if c == 2:
                meas.append('netdata.%s.user' % content.cell(row=1, column=c)._value)
                continue
            meas.append('netdata.%s.used' % content.cell(row=1, column=c)._value)
        measurement_list[content.cell(row=r, column=1)._value] = meas
    print measurement_list

    return measurement_list

if __name__ == '__main__':
    get_env("/Users/k2data/Desktop/env.xlsx")
    get_threshold("/Users/k2data/Desktop/env.xlsx")
    get_host("/Users/k2data/Desktop/env.xlsx")
    get_measurement("/Users/k2data/Desktop/env.xlsx")